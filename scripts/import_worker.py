"""Eligibility & Claims import worker.

Runs on the SQL box (stable IP). Reads active feed configs from dbo.Import_Configs,
pulls matching files from each client's SFTP, maps columns, and inserts into the
configured target table. Designed to be invoked frequently by cron (e.g. every
10-15 min); each config runs only when its schedule says it is due.

Env:
  IRX_DB_PWD         SQL Server password for user 'claudeservices'
  IMPORT_CRYPT_KEY   64 hex chars (32 bytes), SAME value as the Netlify env var

Optional overrides: SQLSERVER_HOST (default 74.117.224.152), SQLSERVER_DB (irx),
  SQLSERVER_USER (claudeservices)

Usage:
  python scripts/import_worker.py                # run all due configs
  python scripts/import_worker.py --config 3     # run one config if due
  python scripts/import_worker.py --config 3 --force   # run now, ignore schedule

Requires: pip install pyodbc paramiko openpyxl cryptography
"""
import argparse
import csv
import fnmatch
import io
import os
import posixpath
import sys
from datetime import date as _date, datetime, timedelta
from decimal import Decimal, InvalidOperation

import pyodbc
import paramiko
from cryptography.hazmat.primitives.ciphers.aead import AESGCM

date_min = _date.min  # sort key for rows whose start date will not parse


# ── DB ──────────────────────────────────────────────────────────────────────
def odbc_driver():
    """Newest installed SQL Server ODBC driver, or SQLSERVER_ODBC_DRIVER if set.
    Pinning to a single version breaks on any box that ships a different one."""
    override = os.environ.get("SQLSERVER_ODBC_DRIVER")
    if override:
        return override
    installed = set(pyodbc.drivers())
    for name in ("ODBC Driver 18 for SQL Server", "ODBC Driver 17 for SQL Server",
                 "ODBC Driver 13.1 for SQL Server", "SQL Server Native Client 11.0"):
        if name in installed:
            return name
    raise RuntimeError(f"No SQL Server ODBC driver found. Installed: {sorted(installed)}")


def db_connect():
    conn = (
        f"DRIVER={{{odbc_driver()}}};"
        f"SERVER={os.environ.get('SQLSERVER_HOST', '74.117.224.152')};"
        f"DATABASE={os.environ.get('SQLSERVER_DB', 'irx')};"
        f"UID={os.environ.get('SQLSERVER_USER', 'claudeservices')};"
        "PWD=" + os.environ["IRX_DB_PWD"] + ";"
        "Encrypt=yes;TrustServerCertificate=yes;"
    )
    return pyodbc.connect(conn, autocommit=True)


def rows_as_dicts(cur):
    cols = [c[0] for c in cur.description]
    return [dict(zip(cols, r)) for r in cur.fetchall()]


# ── Crypto (mirrors netlify/functions/_crypto.js) ───────────────────────────
def decrypt(blob):
    if not blob:
        return None
    parts = str(blob).split(":")
    if len(parts) != 4 or parts[0] != "v1":
        raise ValueError("bad ciphertext format")
    key = bytes.fromhex(os.environ["IMPORT_CRYPT_KEY"])
    iv = bytes.fromhex(parts[1])
    tag = bytes.fromhex(parts[2])
    ct = bytes.fromhex(parts[3])
    return AESGCM(key).decrypt(iv, ct + tag, None).decode("utf-8")


# ── Schedule ────────────────────────────────────────────────────────────────
def is_due(cfg, now):
    freq = cfg["schedule_frequency"]
    last = cfg["last_run_at"]
    if freq == "Hourly":
        return last is None or (now - last) >= timedelta(minutes=55)
    # Daily / Weekly are gated by time-of-day; default 06:00.
    hh, mm = (cfg["schedule_time"] or "06:00").split(":")
    after_time = now.time() >= datetime(now.year, now.month, now.day, int(hh), int(mm)).time()
    ran_today = last is not None and last.date() >= now.date()
    if freq == "Weekly":
        dow = int(now.strftime("%w"))  # 0=Sun..6=Sat, matches schedule_dow
        if cfg["schedule_dow"] is None or dow != int(cfg["schedule_dow"]):
            return False
    return after_time and not ran_today


# ── File parsing ────────────────────────────────────────────────────────────
def _first_cell(r):
    for c in r:
        s = str(c).strip()
        if s:
            return s
    return ""


def split_rows(rows, has_header, header_row, stop_on_blank=False, stop_marker=None, footer_skip=0):
    """Pick the header at the 1-based file row `header_row` (preamble/title rows
    above it are ignored), then bound the data region. Rows are passed in literal
    file order with no prior blank-row filtering so the row number matches Excel.

    End-of-data is the earliest of: first blank row (if stop_on_blank), first row
    whose first value starts with stop_marker, or end of file. footer_skip then
    drops that many trailing rows (totals/notes)."""
    hidx = max(0, (int(header_row) if header_row else 1) - 1)
    if hidx >= len(rows):
        return [], []
    if has_header:
        header = [str(c).strip() for c in rows[hidx]]
        body = rows[hidx + 1:]
    else:
        body = rows[hidx:]
        width = max((len(r) for r in body), default=0)
        header = [str(i + 1) for i in range(width)]

    marker = (stop_marker or "").strip().lower()
    end = len(body)
    for i, r in enumerate(body):
        blank = not any(str(c).strip() for c in r)
        if stop_on_blank and blank:
            end = i
            break
        if marker and _first_cell(r).lower().startswith(marker):
            end = i
            break
    body = body[:end]

    while body and not any(str(c).strip() for c in body[-1]):  # trailing blanks
        body.pop()
    skip = int(footer_skip or 0)
    if skip > 0:
        body = body[:-skip] if skip < len(body) else []
    body = [r for r in body if any(str(c).strip() for c in r)]  # interior blanks
    return header, body


def parse_file(data, cfg):
    opts = dict(stop_on_blank=bool(cfg.get("stop_on_blank")),
                stop_marker=cfg.get("stop_marker"),
                footer_skip=cfg.get("footer_skip") or 0)
    if cfg["file_format"] == "xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = cfg.get("sheet_name")
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = [["" if c is None else c for c in row] for row in ws.iter_rows(values_only=True)]
    else:
        text = data.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(io.StringIO(text), delimiter=(cfg.get("delimiter") or ",")))
    return split_rows(rows, cfg["has_header"], cfg.get("header_row") or 1, **opts)


# ── Value coercion ──────────────────────────────────────────────────────────
def coerce(value, dtype):
    if value is None:
        return None
    s = value if isinstance(value, str) else str(value)
    s = s.strip()
    if s == "":
        return None
    if dtype == "int":
        try:
            return int(float(s))
        except ValueError:
            return None
    if dtype == "decimal":
        try:
            return Decimal(s.replace(",", ""))
        except (InvalidOperation, ValueError):
            return None
    if dtype in ("date", "datetime"):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M"):
            try:
                dt = datetime.strptime(s, fmt)
                return dt.date() if dtype == "date" else dt
            except ValueError:
                continue
        return None
    return s


# ── Import one file ─────────────────────────────────────────────────────────
def import_rows(cur, target_table, maps, header, data_rows, truncate):
    idx = {name: i for i, name in enumerate(header)}
    # Resolve each mapping's source position; warn-skip mappings whose source is absent.
    resolved = []
    for m in maps:
        src = m["source_column"]
        pos = idx.get(src)
        if pos is None and src.isdigit():
            pos = int(src) - 1  # 1-based index fallback
        if pos is None:
            raise ValueError(f"source column '{src}' not found in file header")
        resolved.append((pos, m["target_column"], m.get("data_type")))

    target_cols = [t for _, t, _ in resolved]
    collist = ", ".join(f"[{c}]" for c in target_cols)
    placeholders = ", ".join("?" for _ in target_cols)

    out = []
    for r in data_rows:
        out.append(tuple(
            coerce(r[pos] if pos < len(r) else None, dt) for pos, _, dt in resolved
        ))

    if truncate:
        cur.execute(f"TRUNCATE TABLE {target_table}")
    if out:
        cur.fast_executemany = True
        cur.executemany(f"INSERT INTO {target_table} ({collist}) VALUES ({placeholders})", out)
    return len(out)


# ── Eligibility reconciliation ──────────────────────────────────────────────
def read_stage(cur, stage_table, recon_maps, stage_filter=None):
    """Read the freshly-loaded staging table and project it onto canonical
    eligibility columns using the stage->canonical reconcile map.

    A mapping may carry a `stage_expression` (any SQL expression over the staging
    table) instead of a plain column. Per-client transforms live there -- deriving
    a member key, normalising dates, turning a vendor's placeholder text into NULL
    -- which keeps those rules declarative instead of forking this worker per feed.

    `stage_filter` is an optional WHERE clause used to drop rows that must never
    reach eligibility (e.g. a test record the client ships in every file)."""
    if not recon_maps:
        raise ValueError("no reconcile mapping defined (staging column -> eligibility column)")
    canon_cols = [m["eligibility_column"] for m in recon_maps]
    select_list = ", ".join(
        f"{m.get('stage_expression') or '[' + m['stage_column'] + ']'} AS col{i}"
        for i, m in enumerate(recon_maps))
    where = f" WHERE {stage_filter}" if stage_filter else ""
    cur.execute(f"SELECT {select_list} FROM {stage_table}{where}")
    rows = []
    for r in cur.fetchall():
        d = {}
        for i, m in enumerate(recon_maps):
            v = r[i]
            s = "" if v is None else str(v).strip()
            d[m["eligibility_column"]] = s or None
        rows.append(d)
    return rows, canon_cols


def parse_date_any(s):
    """Parse the date shapes present in dbo.eligibility.

    Historic loads wrote Excel's *display* text rather than a real date, so the
    same column holds '12/31/2039', '39-12-31 0:00' and '00-0-1 12:00' (a null
    date). An unparsed value is read by the caller as "no end date" = still
    active, so failing to understand the yy-m-d shape silently keeps terminated
    members active. Returns None only for genuinely empty/null values."""
    s = str(s).strip()
    if not s:
        return None
    if s.startswith("00-0-") or s.startswith("0000-00-00"):  # Excel's null date
        return None
    s = s.split(" ")[0]  # drop any trailing time component
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%Y%m%d", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # yy-m-d with unpadded month/day, e.g. '39-12-31' -> 2039-12-31.
    try:
        y, m, d = s.split("-")
        if len(y) == 2:
            return datetime(2000 + int(y), int(m), int(d)).date()
    except (ValueError, TypeError):
        pass
    return None


def reconcile_eligibility(cur, cfg, run_id, rows, target_cols):
    """Compare the imported roster to dbo.eligibility for this client (CARRIER =
    irx_client_id), keyed on CARRIER + MEMBER_ID. New members are inserted,
    matched members updated, and active members missing from the file have their
    MEMBER_THRU_DATE set to the run date. Returns (added, updated, inactivated)."""
    table = cfg.get("reconcile_table") or "dbo.eligibility"
    cur.execute("SELECT irx_client_id FROM dbo.tp_clients WHERE id=?", cfg["client_id"])
    row = cur.fetchone()
    carrier = (str(row[0]).strip() if row and row[0] else None)
    if not carrier:
        raise ValueError("client has no irx_client_id; cannot scope eligibility by CARRIER")
    if "MEMBER_ID" not in target_cols:
        raise ValueError("Eligibility feeds must map a column to MEMBER_ID")

    today = datetime.now().date()
    run_date = f"{today.month}/{today.day}/{today.year}"

    # Dedupe file rows by MEMBER_ID, keeping the most recent coverage span.
    # Some feeds (e.g. UOP/WellDyne) carry one row per eligibility span, so a
    # member appears several times -- UOP ships ~2,400 rows for ~1,300 members.
    # Sorting by MEMBER_FROM_DATE first makes "last wins" mean "latest span"
    # rather than "whatever order the file happened to be in". Rows with an
    # unparseable start date sort earliest so a real date always beats them.
    rows = sorted(rows, key=lambda d: parse_date_any(d.get("MEMBER_FROM_DATE") or "") or date_min)
    file_map = {}
    for d in rows:
        mid = (str(d.get("MEMBER_ID")).strip() if d.get("MEMBER_ID") else "")
        if mid:
            file_map[mid] = d

    cur.execute(
        f"SELECT MEMBER_ID, MEMBER_THRU_DATE, LAST_NAME, FIRST_NAME, DATE_OF_BIRTH FROM {table} WHERE CARRIER=?",
        carrier)
    existing = {}
    for r in cur.fetchall():
        existing[(str(r[0]).strip() if r[0] else "")] = {
            "thru": r[1], "last": r[2], "first": r[3], "dob": r[4]}

    insert_cols = ["CARRIER"] + target_cols + ["LoadUpdateDate"]
    update_cols = [c for c in target_cols if c != "MEMBER_ID"]  # never rewrite the key

    inserts, updates, add_items = [], [], []
    for mid, d in file_map.items():
        if mid in existing:
            if update_cols:
                updates.append(tuple([d.get(c) for c in update_cols] + [today, carrier, mid]))
        else:
            inserts.append(tuple([carrier] + [d.get(c) for c in target_cols] + [today]))
            add_items.append((mid, d.get("LAST_NAME"), d.get("FIRST_NAME"), d.get("DATE_OF_BIRTH")))

    inactivations, inact_items = [], []
    for mid, info in existing.items():
        if mid and mid not in file_map:
            thru = parse_date_any(info["thru"])
            if thru is None or thru >= today:  # blank/future = currently active
                inactivations.append((run_date, today, carrier, mid))
                inact_items.append((mid, info["last"], info["first"], info["dob"]))

    cur.fast_executemany = True
    if inserts:
        cols = ", ".join(f"[{c}]" for c in insert_cols)
        ph = ", ".join("?" for _ in insert_cols)
        cur.executemany(f"INSERT INTO {table} ({cols}) VALUES ({ph})", inserts)
    if updates:
        sets = ", ".join(f"[{c}]=?" for c in update_cols) + ", [LoadUpdateDate]=?"
        cur.executemany(f"UPDATE {table} SET {sets} WHERE CARRIER=? AND MEMBER_ID=?", updates)
    if inactivations:
        cur.executemany(
            f"UPDATE {table} SET [MEMBER_THRU_DATE]=?, [LoadUpdateDate]=? WHERE CARRIER=? AND MEMBER_ID=?",
            inactivations)

    item_sql = ("INSERT INTO dbo.Import_Reconcile_Items "
                "(run_id, config_id, action, carrier, member_id, last_name, first_name, date_of_birth) "
                "VALUES (?,?,?,?,?,?,?,?)")
    if add_items:
        cur.executemany(item_sql, [(run_id, cfg["id"], "Add", carrier, *it) for it in add_items])
    if inact_items:
        cur.executemany(item_sql, [(run_id, cfg["id"], "Inactivate", carrier, *it) for it in inact_items])

    return len(inserts), len(updates), len(inactivations)


# ── SFTP ────────────────────────────────────────────────────────────────────
def sftp_connect(cfg):
    transport = paramiko.Transport((cfg["sftp_host"], int(cfg["sftp_port"] or 22)))
    pkey = None
    key_pem = decrypt(cfg.get("sftp_key_enc"))
    if key_pem:
        pkey = paramiko.RSAKey.from_private_key(io.StringIO(key_pem))
    password = decrypt(cfg.get("sftp_password_enc"))
    transport.connect(username=cfg["sftp_username"], password=password, pkey=pkey)
    return paramiko.SFTPClient.from_transport(transport), transport


# ── Run a single config ─────────────────────────────────────────────────────
def run_config(cn, cfg):
    cur = cn.cursor()
    cur.execute(
        "INSERT INTO dbo.Import_Runs (config_id, status) OUTPUT INSERTED.id VALUES (?, 'Running')",
        cfg["id"])
    run_id = cur.fetchone()[0]

    def finish(status, file_name=None, rows=None, message=None):
        cur.execute(
            "UPDATE dbo.Import_Runs SET finished_at=GETDATE(), status=?, file_name=?, rows_imported=?, message=? WHERE id=?",
            status, file_name, rows, (message[:3900] if message else None), run_id)
        cur.execute("UPDATE dbo.Import_Configs SET last_run_at=GETDATE() WHERE id=?", cfg["id"])

    try:
        cur.execute(
            "SELECT source_column, target_column, data_type FROM dbo.Import_Column_Maps WHERE config_id=? ORDER BY ordinal, id",
            cfg["id"])
        maps = rows_as_dicts(cur)
        if not maps:
            finish("Error", message="No column mappings defined")
            return

        sftp, transport = sftp_connect(cfg)
        try:
            listing = sftp.listdir_attr(cfg["remote_dir"] or "/")
        finally:
            pass

        names = [a.filename for a in listing
                 if fnmatch.fnmatch(a.filename, cfg["file_pattern"] or "*")]
        cur.execute("SELECT file_name FROM dbo.Import_Processed_Files WHERE config_id=?", cfg["id"])
        done = {r[0] for r in cur.fetchall()}
        todo = sorted(n for n in names if n not in done)

        if not todo:
            sftp.close(); transport.close()
            finish("NoFile", message=f"No new files matching {cfg['file_pattern']}")
            return

        # Eligibility = two-stage: raw-load into the staging (target) table, then
        # reconcile that table into the canonical reconcile_table.
        if cfg["feed_type"] == "Eligibility":
            cur.execute(
                "SELECT stage_column, eligibility_column, stage_expression FROM dbo.Import_Reconcile_Maps WHERE config_id=? ORDER BY ordinal, id",
                cfg["id"])
            recon_maps = rows_as_dicts(cur)
            if not recon_maps:
                sftp.close(); transport.close()
                finish("Error", message="No reconcile mapping (staging column -> eligibility column) defined")
                return

            # Stage 1: raw-load every new file into the staging table (full refresh).
            staged = 0
            for i, name in enumerate(todo):
                remote_path = posixpath.join(cfg["remote_dir"] or "/", name)
                with sftp.open(remote_path, "rb") as fh:
                    data = fh.read()
                header, body = parse_file(data, cfg)
                staged += import_rows(cur, cfg["target_table"], maps, header, body, truncate=(i == 0))
                cur.execute(
                    "INSERT INTO dbo.Import_Processed_Files (config_id, file_name, rows_imported) VALUES (?,?,?)",
                    cfg["id"], name, staged if i == len(todo) - 1 else 0)
                if cfg["after_import"] == "delete":
                    sftp.remove(remote_path)
                elif cfg["after_import"] == "archive" and cfg.get("archive_dir"):
                    try:
                        sftp.rename(remote_path, posixpath.join(cfg["archive_dir"], name))
                    except IOError:
                        pass
            sftp.close(); transport.close()

            # Stage 2: reconcile staging -> canonical eligibility.
            rows, target_cols = read_stage(cur, cfg["target_table"], recon_maps,
                                           cfg.get("stage_filter"))
            added, updated, inactivated = reconcile_eligibility(cur, cfg, run_id, rows, target_cols)
            cur.execute(
                "UPDATE dbo.Import_Runs SET added_count=?, updated_count=?, inactivated_count=? WHERE id=?",
                added, updated, inactivated, run_id)
            finish("Success", file_name=(todo[0] if len(todo) == 1 else f"{len(todo)} files"),
                   rows=staged,
                   message=f"Staged {staged} rows -> {cfg.get('reconcile_table') or 'dbo.eligibility'}: "
                           f"{added} added, {updated} updated, {inactivated} inactivated")
            return

        total = 0
        last_file = None
        for name in todo:
            remote_path = posixpath.join(cfg["remote_dir"] or "/", name)
            with sftp.open(remote_path, "rb") as fh:
                data = fh.read()
            header, body = parse_file(data, cfg)
            n = import_rows(cur, cfg["target_table"], maps, header, body,
                            cfg["truncate_before"] and name == todo[0])
            cur.execute(
                "INSERT INTO dbo.Import_Processed_Files (config_id, file_name, rows_imported) VALUES (?,?,?)",
                cfg["id"], name, n)
            total += n
            last_file = name
            # after-import disposition
            if cfg["after_import"] == "delete":
                sftp.remove(remote_path)
            elif cfg["after_import"] == "archive" and cfg.get("archive_dir"):
                try:
                    sftp.rename(remote_path, posixpath.join(cfg["archive_dir"], name))
                except IOError:
                    pass

        sftp.close(); transport.close()
        finish("Success", file_name=(last_file if len(todo) == 1 else f"{len(todo)} files"),
               rows=total, message=f"Imported {total} rows from {len(todo)} file(s)")
    except Exception as e:  # noqa: BLE001 - record any failure on the run
        finish("Error", message=f"{type(e).__name__}: {e}")
        print(f"[config {cfg['id']}] ERROR: {e}", file=sys.stderr)


# ── Main ────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", type=int, help="run only this config id")
    ap.add_argument("--force", action="store_true", help="ignore schedule (with --config)")
    args = ap.parse_args()

    cn = db_connect()
    cur = cn.cursor()
    # Active configs (for scheduled runs) plus any flagged for a manual run.
    where = "WHERE (active=1 OR run_requested=1)" + (" AND id=?" if args.config else "")
    params = (args.config,) if args.config else ()
    cur.execute(
        "SELECT id, client_id, name, feed_type, sftp_host, sftp_port, sftp_username, "
        "sftp_password_enc, sftp_key_enc, remote_dir, file_pattern, file_format, delimiter, "
        "has_header, header_row, stop_on_blank, stop_marker, footer_skip, sheet_name, "
        "target_table, reconcile_table, stage_filter, truncate_before, after_import, archive_dir, "
        "schedule_frequency, schedule_time, schedule_dow, active, run_requested, last_run_at "
        f"FROM dbo.Import_Configs {where}", *params)
    cfgs = rows_as_dicts(cur)

    now = datetime.now()
    ran = 0
    for cfg in cfgs:
        requested = bool(cfg.get("run_requested"))
        due_scheduled = bool(cfg.get("active")) and is_due(cfg, now)
        if args.force or requested or due_scheduled:
            reason = "manual" if requested else ("forced" if args.force else "scheduled")
            print(f"Running config {cfg['id']} ({cfg['name']}) [{reason}]…")
            run_config(cn, cfg)
            ran += 1
        if requested:  # clear the manual-run flag whether or not it just ran
            cur.execute("UPDATE dbo.Import_Configs SET run_requested=0 WHERE id=?", cfg["id"])
    print(f"Done. {ran} of {len(cfgs)} config(s) run.")
    cn.close()


if __name__ == "__main__":
    main()
