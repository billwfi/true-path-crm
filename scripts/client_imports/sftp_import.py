"""Client claims/eligibility SFTP importer (Step 1: get new data in).

Rerunnable, repo-tracked setup for pulling a client's eligibility + claims files
from SFTP into per-client SQL Server staging tables. Each run TRUNCATES the target
table and reloads it, so the table always holds the latest file. A later Step 2
(separate script) compares these staging tables to the production eligibility /
claims tables and emails a reconciliation report.

Add a client by appending an entry to CLIENTS below — no code changes needed.

Usage:
  python scripts/client_imports/sftp_import.py <client>            # all feeds
  python scripts/client_imports/sftp_import.py <client> <feed>     # one feed by name
  python scripts/client_imports/sftp_import.py mcrhotels --recreate  # DROP+CREATE (schema changed)

Env vars (never commit secrets):
  IRX_DB_PWD      SQL Server password for 'claudeservices' (same as the other scripts)
  MCR_SFTP_PWD    SFTP password for the MCR Hotels feed (see CLIENTS[*]['sftp_pwd_env'])

Requires: pip install pyodbc paramiko openpyxl
"""
import argparse
import fnmatch
import io
import os
import re
import sys
from datetime import datetime, date

import pyodbc
import paramiko
import openpyxl


# ── Client / feed registry ───────────────────────────────────────────────────
# Each feed pulls the newest file matching `pattern` from `remote_dir` and loads
# it into `table` (truncate + reload). `computed` adds derived columns not in the
# file: { new_col: (source_header, fn) } where fn(cell_text) -> value.
CLIENTS = {
    "mcrhotels": {
        "label": "MCR Hotels",
        "client_id": 23,                  # tp_clients.id — links the run log to the client page
        "sftp_host": "us-east-1.sftpcloud.io",
        "sftp_port": 22,
        "sftp_user": "MANAGER",           # case-sensitive
        "sftp_pwd_env": "MCR_SFTP_PWD",
        "remote_dir": "/InternationalRx/MCRHotels",  # case-sensitive path
        "feeds": [
            {
                "name": "Eligibility",
                "pattern": "MCR_Member*.xlsx",
                "table": "Eligibility_MCRHotels",
                "sheet": "Detail",     # roster is on the 'Detail' sheet, not the 'Cover' sheet
                "computed": {},
            },
            {
                "name": "Claims",
                "pattern": "MCRINVESTORS_*.xlsx",
                "table": "ClaimsData_MCRHotels",
                # groupid = SUBSTRING([group id], 1, 8)
                "computed": {"groupid": ("group id", lambda v: (v or "")[:8] or None)},
            },
        ],
    },
}


# ── DB ────────────────────────────────────────────────────────────────────────
def db_connect(autocommit=False):
    conn = (
        "DRIVER={ODBC Driver 17 for SQL Server};"
        f"SERVER={os.environ.get('SQLSERVER_HOST', '74.117.224.152')};"
        f"DATABASE={os.environ.get('SQLSERVER_DB', 'irx')};"
        f"UID={os.environ.get('SQLSERVER_USER', 'claudeservices')};"
        "PWD=" + os.environ["IRX_DB_PWD"] + ";"
        "Encrypt=yes;TrustServerCertificate=yes;"
    )
    return pyodbc.connect(conn, autocommit=autocommit)


def resolve_group(cur, client_id):
    """(group_id, group_name) from tp_clients — GroupID is irx_client_id (CARRIER)."""
    if not client_id:
        return None, None
    row = cur.execute(
        "SELECT irx_client_id, name FROM dbo.tp_clients WHERE id = ?", client_id).fetchone()
    return (row[0], row[1]) if row else (None, None)


# ── Run log (dbo.Client_Import_Log) ──────────────────────────────────────────
def log_start(logcur, client_key, cfg, group_id, group_name, feed, file_name):
    return logcur.execute(
        "INSERT INTO dbo.Client_Import_Log "
        "(client_key, client_id, group_id, group_name, feed_name, target_table, file_name, status) "
        "OUTPUT INSERTED.id VALUES (?,?,?,?,?,?,?, 'Running')",
        client_key, cfg.get("client_id"), group_id, group_name,
        feed["name"], feed["table"], file_name).fetchone()[0]


def log_finish(logcur, log_id, status, rows=None, message=None):
    logcur.execute(
        "UPDATE dbo.Client_Import_Log SET status=?, rows_loaded=?, finished_at=GETDATE(), message=? WHERE id=?",
        status, rows, (message[:3900] if message else None), log_id)


# ── SFTP ──────────────────────────────────────────────────────────────────────
def sftp_connect(cfg):
    pwd = os.environ[cfg["sftp_pwd_env"]]
    transport = paramiko.Transport((cfg["sftp_host"], int(cfg.get("sftp_port", 22))))
    transport.connect(username=cfg["sftp_user"], password=pwd)
    return paramiko.SFTPClient.from_transport(transport), transport


def newest_match(sftp, remote_dir, pattern):
    """Return (filename, bytes) of the most recently modified file matching pattern."""
    attrs = [a for a in sftp.listdir_attr(remote_dir)
             if fnmatch.fnmatch(a.filename, pattern)]
    if not attrs:
        return None, None
    attrs.sort(key=lambda a: a.st_mtime or 0, reverse=True)
    name = attrs[0].filename
    with sftp.open(remote_dir.rstrip("/") + "/" + name, "rb") as fh:
        return name, fh.read()


# ── XLSX parsing ──────────────────────────────────────────────────────────────
def parse_xlsx(data, sheet=None):
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    while rows and not any(str(c).strip() for c in rows[-1] if c is not None):
        rows.pop()  # trailing blank rows
    if not rows:
        return [], []
    header = [to_text(c) or "" for c in rows[0]]
    body = [r for r in rows[1:] if any(c is not None and str(c).strip() for c in r)]
    return header, body


def to_text(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return "1" if v else "0"
    if isinstance(v, datetime):
        return v.strftime("%m/%d/%Y") if (v.hour == v.minute == v.second == 0) \
            else v.strftime("%m/%d/%Y %H:%M:%S")
    if isinstance(v, date):
        return v.strftime("%m/%d/%Y")
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else repr(v)
    s = str(v).strip()
    return s if s != "" else None


# ── Column names ──────────────────────────────────────────────────────────────
def sanitize(name):
    s = re.sub(r"[^0-9A-Za-z_]", "_", (name or "").strip())
    s = re.sub(r"_+", "_", s).strip("_")
    if not s:
        s = "col"
    if s[0].isdigit():
        s = "_" + s
    return s


def unique_names(headers):
    """Sanitised, de-duplicated column names, parallel to the header list."""
    out, seen = [], {}
    for h in headers:
        base = sanitize(h)
        n = seen.get(base.lower(), 0) + 1
        seen[base.lower()] = n
        out.append(base if n == 1 else f"{base}_{n}")
    return out


def col_type(values):
    maxlen = max((len(v) for v in values if v is not None), default=0)
    if maxlen > 4000:
        return "NVARCHAR(MAX)"
    return f"NVARCHAR({max(50, min(4000, maxlen * 2 or 50))})"


# ── Load one feed ─────────────────────────────────────────────────────────────
def load_feed(cur, cfg, feed, name, data, recreate):
    header, body = parse_xlsx(data, feed.get("sheet"))
    if not header:
        print(f"  [{feed['name']}] {name}: no rows, skipped")
        return 0

    cols = unique_names(header)
    # Resolve computed source columns against the (case-insensitive) file header.
    hdr_idx = {h.strip().lower(): i for i, h in enumerate(header)}
    computed = []  # (new_col, source_index_or_None, fn)
    for new_col, (src_header, fn) in (feed.get("computed") or {}).items():
        idx = hdr_idx.get(src_header.strip().lower())
        if idx is None:
            raise ValueError(
                f"computed column '{new_col}' needs source header '{src_header}', "
                f"not found in file. Headers: {header}")
        computed.append((sanitize(new_col), idx, fn))

    # Build every output row (file cells as text + computed columns).
    out_rows = []
    for r in body:
        cells = [to_text(r[i]) if i < len(r) else None for i in range(len(header))]
        extra = [fn(cells[idx]) for _, idx, fn in computed]
        out_rows.append(cells + extra)

    all_cols = cols + [c for c, _, _ in computed]
    # Column sizing from the actual data in each position.
    types = []
    for j in range(len(all_cols)):
        types.append(col_type([row[j] for row in out_rows]))

    table = feed["table"]
    qtable = f"dbo.[{table}]"
    exists = cur.execute(
        "SELECT OBJECT_ID(?, 'U')", f"dbo.{table}").fetchone()[0] is not None

    if recreate and exists:
        cur.execute(f"DROP TABLE {qtable}")
        exists = False
    if not exists:
        coldefs = ",\n  ".join(f"[{c}] {t} NULL" for c, t in zip(all_cols, types))
        cur.execute(f"CREATE TABLE {qtable} (\n  {coldefs}\n)")
        print(f"  [{feed['name']}] created {qtable} ({len(all_cols)} columns)")
    else:
        cur.execute(f"TRUNCATE TABLE {qtable}")

    collist = ", ".join(f"[{c}]" for c in all_cols)
    placeholders = ", ".join("?" for _ in all_cols)
    cur.fast_executemany = True
    if out_rows:
        cur.executemany(
            f"INSERT INTO {qtable} ({collist}) VALUES ({placeholders})", out_rows)
    print(f"  [{feed['name']}] {name}: loaded {len(out_rows)} rows into {qtable}")
    return len(out_rows)


# ── Run a client ──────────────────────────────────────────────────────────────
def run_client(client_key, only_feed=None, recreate=False):
    cfg = CLIENTS.get(client_key)
    if not cfg:
        sys.exit(f"Unknown client '{client_key}'. Known: {', '.join(CLIENTS)}")
    for var in ("IRX_DB_PWD", cfg["sftp_pwd_env"]):
        if not os.environ.get(var):
            sys.exit(f"Missing required env var: {var}")

    print(f"== {cfg['label']} ({client_key}) — {cfg['remote_dir']} ==")
    sftp, transport = sftp_connect(cfg)
    cn = db_connect()                 # transactional: one commit per feed
    logcn = db_connect(autocommit=True)  # run log persists even if a feed fails
    cur, logcur = cn.cursor(), logcn.cursor()
    group_id, group_name = resolve_group(cur, cfg.get("client_id"))
    try:
        total = 0
        for feed in cfg["feeds"]:
            if only_feed and feed["name"].lower() != only_feed.lower():
                continue
            name, data = newest_match(sftp, cfg["remote_dir"], feed["pattern"])
            log_id = log_start(logcur, client_key, cfg, group_id, group_name, feed, name)
            if not name:
                print(f"  [{feed['name']}] no file matching {feed['pattern']}")
                log_finish(logcur, log_id, "NoFile", message=f"no file matching {feed['pattern']}")
                continue
            try:
                n = load_feed(cur, cfg, feed, name, data, recreate)
                cn.commit()
                log_finish(logcur, log_id, "Success", rows=n,
                           message=f"loaded {n} rows into dbo.{feed['table']}")
                total += n
            except Exception as e:  # record the failure, keep going with other feeds
                cn.rollback()
                log_finish(logcur, log_id, "Error", message=f"{type(e).__name__}: {e}")
                print(f"  [{feed['name']}] ERROR: {e}", file=sys.stderr)
        print(f"Done. {total} rows loaded.")
    finally:
        cur.close(); cn.close(); logcur.close(); logcn.close()
        sftp.close(); transport.close()


def main():
    ap = argparse.ArgumentParser(description="Client SFTP claims/eligibility importer")
    ap.add_argument("client", help="client key, e.g. mcrhotels")
    ap.add_argument("feed", nargs="?", help="optional single feed name (Eligibility | Claims)")
    ap.add_argument("--recreate", action="store_true",
                    help="DROP and recreate the table instead of truncating (use when the file schema changed)")
    args = ap.parse_args()
    run_client(args.client, args.feed, args.recreate)


if __name__ == "__main__":
    main()
